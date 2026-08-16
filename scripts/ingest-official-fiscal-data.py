#!/usr/bin/env python3
"""Ingest official MDIC tariff and RFB TIPI workbooks into a candidate snapshot.

The official workbooks have changed layout over time.  This parser therefore uses
header detection when available and a data-driven column inference fallback.  It
also normalizes Unicode/accents and preserves every unambiguous NCM/rate row.
"""
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

NCM_RE = re.compile(r"^\d{8}$")


def text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def norm_key(v: Any) -> str:
    raw = unicodedata.normalize("NFKD", text(v))
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", raw.lower())


def normalize_ncm(v: Any) -> str | None:
    raw = text(v).replace(".", "").replace("-", "").replace(" ", "")
    if raw.endswith(".0"):
        raw = raw[:-2]
    if raw.isdigit() and len(raw) == 7:
        raw = "0" + raw
    return raw if NCM_RE.fullmatch(raw) else None


def parse_rate(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        rate = float(v)
        return rate if 0 <= rate <= 100 else None
    raw = text(v).replace("%", "").replace(" ", "")
    if not raw:
        return None
    if raw.count(",") == 1:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        rate = float(raw)
    except ValueError:
        return None
    return rate if 0 <= rate <= 100 else None


def find_col(headers: list[Any], exact: set[str], contains: tuple[str, ...]) -> int | None:
    keys = [norm_key(x) for x in headers]
    for candidate in exact:
        if candidate in keys:
            return keys.index(candidate)
    for i, key in enumerate(keys):
        if any(part in key for part in contains):
            return i
    return None


def find_header(rows: list[tuple[Any, ...]], rate_exact: set[str], rate_contains: tuple[str, ...]):
    for idx, row in enumerate(rows[:100]):
        headers = list(row)
        ncm_col = find_col(
            headers,
            {"ncm", "codigo", "codigoncm", "codigoncmsh", "codigoncm2022"},
            ("ncm", "codigo"),
        )
        rate_col = find_col(headers, rate_exact, rate_contains)
        if ncm_col is not None and rate_col is not None and ncm_col != rate_col:
            return idx, ncm_col, rate_col
    return None


def infer_columns(rows: list[tuple[Any, ...]], source_type: str) -> tuple[int | None, int | None]:
    """Infer NCM/rate columns from actual data when the workbook has no usable header."""
    if not rows:
        return None, None
    width = max(len(r) for r in rows)
    sample = rows[: min(len(rows), 1500)]
    ncm_scores = [0] * width
    rate_scores = [0] * width

    for row in sample:
        for col in range(len(row)):
            value = row[col]
            if normalize_ncm(value) is not None:
                ncm_scores[col] += 1
            if parse_rate(value) is not None:
                rate_scores[col] += 1

    ncm_col = max(range(width), key=ncm_scores.__getitem__) if width else None
    if ncm_col is None or ncm_scores[ncm_col] < 2:
        return None, None

    # The rate column should contain many numeric/rate values but must not be the NCM column.
    candidates = [c for c in range(width) if c != ncm_col]
    if not candidates:
        return None, None
    rate_col = max(candidates, key=rate_scores.__getitem__)
    if rate_scores[rate_col] < 2:
        return ncm_col, None

    # Avoid selecting a description/quantity column with lots of arbitrary numbers.
    # Prefer a rate column that appears adjacent to NCM when scores are comparable.
    best_score = rate_scores[rate_col]
    nearby = [c for c in candidates if abs(c - ncm_col) <= 4]
    for c in nearby:
        if rate_scores[c] >= max(2, int(best_score * 0.70)):
            rate_col = c
            break
    return ncm_col, rate_col


def extract(path: Path, source_type: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    if source_type == "mdic-ii":
        exact, contains = {
            "aliquota", "aliquotaii", "ii", "tarifa", "tec", "aliquotatec", "impostoimportacao"
        }, ("aliquota", "tarifa", "tec", "importacao")
    else:
        exact, contains = {"ipi", "aliquotaipi", "aliquotai", "aliquota"}, ("ipi", "aliquota")

    records: list[dict[str, Any]] = []
    rejected = 0
    sheets = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        sheets.append({
            "name": sheet.title,
            "rows": len(rows),
            "cols": max((len(r) for r in rows), default=0),
        })
        if not rows:
            continue

        header = find_header(rows, exact, contains)
        if header is not None:
            hidx, ncm_col, rate_col = header
        else:
            hidx = 0
            ncm_col, rate_col = infer_columns(rows, source_type)

        if ncm_col is None or rate_col is None:
            print(f"SKIP sheet={sheet.title!r}: unable to infer NCM/rate columns")
            continue

        print(f"PARSE sheet={sheet.title!r}: header={header is not None} ncm_col={ncm_col + 1} rate_col={rate_col + 1}")
        start = hidx + 1 if header is not None else 0
        for row_number, row in enumerate(rows[start:], start=start + 1):
            if not any(v not in (None, "") for v in row):
                continue
            ncm = normalize_ncm(row[ncm_col] if ncm_col < len(row) else None)
            rate = parse_rate(row[rate_col] if rate_col < len(row) else None)
            if ncm is None or rate is None:
                rejected += 1
                continue
            records.append({
                "sourceType": source_type,
                "ncm": ncm,
                "rate": rate,
                "sheet": sheet.title,
                "row": row_number,
                "workbook": path.name,
            })

    return records, rejected, sheets


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--mdic", required=True, type=Path)
    p.add_argument("--tipi", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    a = p.parse_args()

    mdic, mdic_rej, mdic_sheets = extract(a.mdic, "mdic-ii")
    tipi, tipi_rej, tipi_sheets = extract(a.tipi, "rfb-ipi")
    print(f"MDIC sheets={len(mdic_sheets)} records={len(mdic)} rejected={mdic_rej}")
    print(f"TIPI sheets={len(tipi_sheets)} records={len(tipi)} rejected={tipi_rej}")

    if not mdic:
        raise SystemExit("MDIC ingestion produced zero records; refusing publication.")
    if not tipi:
        raise SystemExit("TIPI ingestion produced zero records; refusing publication.")

    out = {
        "schemaVersion": 3,
        "publicationStatus": "candidate",
        "sources": {
            "mdic": {
                "file": a.mdic.name,
                "published": "2026-07-24",
                "recordCount": len(mdic),
                "rejectedRows": mdic_rej,
                "sheets": mdic_sheets,
            },
            "rfbTipi": {
                "file": a.tipi.name,
                "updated": "2026-02-13",
                "recordCount": len(tipi),
                "rejectedRows": tipi_rej,
                "sheets": tipi_sheets,
            },
        },
        "records": mdic + tipi,
        "notes": [
            "All unambiguous source rows are preserved; duplicate NCMs are intentional across tariff treatments.",
            "MDIC annex precedence is resolved by the fiscal engine.",
            "Snapshot remains candidate until acceptance tests pass.",
        ],
    }
    a.output.parent.mkdir(parents=True, exist_ok=True)
    a.output.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated {len(out['records'])} total records")


if __name__ == "__main__":
    main()
