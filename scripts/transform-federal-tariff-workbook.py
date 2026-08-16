#!/usr/bin/env python3
"""Transform the official MDIC tariff workbook into the ImportaFácil catalog contract.

Conservative by design: it normalizes NCMs, extracts candidate II rates,
preserves provenance, rejects ambiguous rows, detects conflicting duplicates,
and never publishes an empty catalog. The generated snapshot remains a
candidate until source-specific legal precedence is reviewed.

Usage:
  python scripts/transform-federal-tariff-workbook.py input.xlsx output.json
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

NCM_RE = re.compile(r"^\d{8}$")


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_ncm(value: Any) -> str | None:
    value = text(value).replace(".", "").replace(" ", "").replace("-", "")
    if value.isdigit() and len(value) == 7:
        value = "0" + value
    return value if NCM_RE.fullmatch(value) else None


def key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def find_column(headers: list[Any], candidates: set[str]) -> int | None:
    normalized = [key(header) for header in headers]
    for candidate in candidates:
        if candidate in normalized:
            return normalized.index(candidate)
    return None


def parse_rate(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace("%", "").replace(" ", "")
    if raw.count(",") == 1:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[Any]] | None:
    ncm_candidates = {"ncm", "codigo", "codigoncm", "codigoncmsh", "codigoncm2022"}
    rate_candidates = {"aliquota", "aliquotaii", "ii", "tarifa", "tec", "aliquotatec"}
    for index, row in enumerate(rows[:25]):
        if find_column(list(row), ncm_candidates) is not None and find_column(list(row), rate_candidates) is not None:
            return index, list(row)
    return None


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/transform-federal-tariff-workbook.py <input.xlsx> <output.json>")

    source, destination = map(Path, sys.argv[1:])
    if not source.exists():
        raise SystemExit(f"Workbook not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    seen: dict[str, dict[str, Any]] = {}
    conflicts: list[dict[str, Any]] = []

    ncm_candidates = {"ncm", "codigo", "codigoncm", "codigoncmsh", "codigoncm2022"}
    rate_candidates = {"aliquota", "aliquotaii", "ii", "tarifa", "tec", "aliquotatec"}

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header = find_header_row(rows)
        if header is None:
            continue
        header_index, headers = header
        ncm_col = find_column(headers, ncm_candidates)
        rate_col = find_column(headers, rate_candidates)
        assert ncm_col is not None and rate_col is not None

        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(value not in (None, "") for value in row):
                continue
            ncm = normalize_ncm(row[ncm_col] if ncm_col < len(row) else None)
            rate = parse_rate(row[rate_col] if rate_col < len(row) else None)
            if ncm is None or rate is None or rate < 0 or rate > 100:
                rejected.append({"sheet": sheet.title, "row": row_number, "reason": "ambiguous_ncm_or_rate"})
                continue

            record = {
                "ncm": ncm,
                "iiRate": rate,
                "sourceSheet": sheet.title,
                "sourceRow": row_number,
                "sourceWorkbook": source.name,
                "treatment": "official-mdic-workbook",
            }
            previous = seen.get(ncm)
            if previous is None:
                seen[ncm] = record
                records.append(record)
            elif previous["iiRate"] != rate:
                conflicts.append({"ncm": ncm, "first": previous, "second": record})

    if not records:
        raise SystemExit("No unambiguous NCM/rate records found; refusing to generate an empty catalog.")
    if conflicts:
        raise SystemExit(f"Conflicting tariff records found for {len(conflicts)} NCM(s); refusing to publish ambiguous data.")

    output = {
        "schemaVersion": 2,
        "source": source.name,
        "recordCount": len(records),
        "rejectedCount": len(rejected),
        "conflictCount": len(conflicts),
        "records": records,
        "rejected": rejected,
        "publicationStatus": "candidate",
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"generated {len(records)} candidate records; rejected {len(rejected)} rows; conflicts {len(conflicts)}")


if __name__ == "__main__":
    main()
