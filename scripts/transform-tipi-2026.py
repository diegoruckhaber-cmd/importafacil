#!/usr/bin/env python3
"""Transform the official Receita Federal TIPI XLSX into a candidate catalog.

The TIPI workbook format has changed over time, so this loader discovers likely
NCM and IPI columns instead of assuming fixed positions. Ambiguous rows are
rejected and no application data is mutated by the script.

Usage:
  python scripts/transform-tipi-2026.py input.xlsx output.json
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

NCM_RE = re.compile(r"^\d{8}$")


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", text(value).lower())


def normalize_ncm(value: Any) -> str | None:
    raw = text(value).replace(".", "").replace(" ", "")
    return raw if NCM_RE.fullmatch(raw) else None


def find_column(headers: list[Any], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if key(header) in candidates:
            return index
    return None


def rate(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = text(value).replace("%", "").replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python scripts/transform-tipi-2026.py <input.xlsx> <output.json>")

    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    if not source.exists():
        raise SystemExit(f"Workbook not found: {source}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    ncm_candidates = {"ncm", "codigo", "codigoncm"}
    ipi_candidates = {"ipi", "aliquota", "aliquotaiipi", "aliquotatipi"}

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows, ()))
        ncm_index = find_column(headers, ncm_candidates)
        ipi_index = find_column(headers, ipi_candidates)
        if ncm_index is None:
            continue

        for row_number, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            ncm = normalize_ncm(row[ncm_index] if ncm_index < len(row) else None)
            ipi = rate(row[ipi_index]) if ipi_index is not None and ipi_index < len(row) else None
            if ncm is None or ipi is None:
                rejected.append({"sheet": sheet.title, "row": row_number, "reason": "missing_or_ambiguous_ncm_or_ipi"})
                continue
            records.append({
                "ncm": ncm,
                "ipiRate": ipi,
                "sourceSheet": sheet.title,
                "sourceRow": row_number,
                "sourceWorkbook": source.name,
                "treatment": "official-receita-tipi",
            })

    if not records:
        raise SystemExit("No unambiguous NCM/IPI records found; refusing to generate an empty catalog.")

    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps({
        "schemaVersion": 1,
        "source": source.name,
        "recordCount": len(records),
        "rejectedCount": len(rejected),
        "records": records,
        "rejected": rejected,
        "publicationStatus": "candidate",
        "note": "Candidate snapshot only; validate workbook structure and legal exceptions before production publication.",
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"generated {len(records)} candidate TIPI records; rejected {len(rejected)} rows")


if __name__ == "__main__":
    main()
