#!/usr/bin/env python3
"""Inspect the official MDIC tariff workbook before importing it.

Usage:
  python scripts/inspect-federal-tariff-workbook.py /path/to/tarifas-vigentes.xlsx

The script deliberately does not mutate application data. It reports sheet names,
column headers, row counts and a small sample so the loader can be mapped to the
actual official workbook version before data is committed to the application.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python scripts/inspect-federal-tariff-workbook.py <xlsx>")

    path = Path(sys.argv[1])
    if not path.exists():
        raise SystemExit(f"Workbook not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    report = {"file": str(path), "sheets": []}

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, ())
        row_count = 0
        sample = []
        for row in rows:
            if any(value not in (None, "") for value in row):
                row_count += 1
                if len(sample) < 3:
                    sample.append(list(row))

        report["sheets"].append({
            "name": sheet.title,
            "max_columns": sheet.max_column,
            "non_empty_rows_after_header": row_count,
            "header": list(header),
            "sample": sample,
        })

    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
